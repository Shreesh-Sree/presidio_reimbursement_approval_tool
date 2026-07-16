"""Finance-owned reimbursement payment lifecycle and export operations.

The core reimbursement database deliberately stores no bank-account details.
Payment records carry only report amounts plus opaque processor references; a
future payment rail can resolve a recipient outside this bounded context.
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import UTC, date, datetime
from decimal import Decimal
from typing import Any, Iterable

from sqlalchemy.orm import Session

from app.models.expense_report import ExpenseReport
from app.models.payment_batch import PaymentBatch
from app.models.payment_event import PaymentEvent
from app.models.payment_record import PaymentRecord
from app.models.user import User
from app.services.audit_service import record_audit
from app.services.notification_service import notify


PAYMENT_STATUSES = frozenset({"pending", "batched", "exported", "paid", "failed"})
BATCH_STATUSES = frozenset({"created", "exported"})


class PaymentError(ValueError):
    """A domain error that can safely be returned from the payments API."""


class PaymentNotFoundError(PaymentError):
    pass


class PaymentTransitionError(PaymentError):
    pass


class PaymentValidationError(PaymentError):
    pass


def utcnow() -> datetime:
    return datetime.now(UTC)


def _as_uuid(value: uuid.UUID | str, *, field_name: str) -> uuid.UUID:
    if isinstance(value, uuid.UUID):
        return value
    try:
        return uuid.UUID(str(value))
    except (TypeError, ValueError, AttributeError) as exc:
        raise PaymentValidationError(f"Invalid {field_name}") from exc


def _money(value: Decimal | float | int) -> Decimal:
    return Decimal(str(value)).quantize(Decimal("0.01"))


def _scoped_payment_query(db: Session, payment_id: uuid.UUID, organization_id: uuid.UUID):
    return (
        db.query(PaymentRecord)
        .join(ExpenseReport, ExpenseReport.id == PaymentRecord.expense_report_id)
        .join(User, User.id == ExpenseReport.employee_user_id)
        .filter(
            PaymentRecord.id == payment_id,
            PaymentRecord.is_deleted.is_(False),
            ExpenseReport.is_deleted.is_(False),
            User.is_deleted.is_(False),
            User.organization_id == organization_id,
        )
    )


def _require_actor(db: Session, actor_user_id: uuid.UUID | str, organization_id: uuid.UUID) -> uuid.UUID:
    actor_id = _as_uuid(actor_user_id, field_name="actor user id")
    actor = (
        db.query(User)
        .filter(
            User.id == actor_id,
            User.organization_id == organization_id,
            User.is_deleted.is_(False),
            User.status == "active",
        )
        .first()
    )
    if actor is None:
        raise PaymentValidationError("Payment actor must be an active user in this organization")
    return actor_id


def _payment_report(db: Session, payment: PaymentRecord) -> ExpenseReport:
    report = (
        db.query(ExpenseReport)
        .filter(ExpenseReport.id == payment.expense_report_id, ExpenseReport.is_deleted.is_(False))
        .first()
    )
    if report is None:
        raise PaymentNotFoundError("Expense report not found")
    return report


def _event(
    db: Session,
    payment: PaymentRecord,
    *,
    event_type: str,
    from_status: str | None,
    to_status: str,
    performed_by: uuid.UUID | str | None,
    remarks: str | None = None,
    provider_reference: str | None = None,
) -> PaymentEvent:
    event = PaymentEvent(
        payment_record_id=payment.id,
        payment_batch_id=payment.payment_batch_id,
        event_type=event_type,
        from_status=from_status,
        to_status=to_status,
        amount=_money(payment.amount),
        provider_reference=provider_reference,
        remarks=remarks,
        performed_by=_as_uuid(performed_by, field_name="actor user id") if performed_by else None,
        occurred_at=utcnow(),
    )
    db.add(event)
    return event


def _payment_audit_payload(payment: PaymentRecord) -> dict[str, Any]:
    """Audit only non-sensitive lifecycle values; bank data is never handled."""

    return {
        "status": payment.status,
        "payment_reference": payment.payment_reference,
        "payment_batch_id": str(payment.payment_batch_id) if payment.payment_batch_id else None,
        "provider_reference": payment.provider_reference,
        "payment_date": payment.payment_date.isoformat() if payment.payment_date else None,
    }


def ensure_pending_payment(
    db: Session,
    report: ExpenseReport,
    *,
    created_by: uuid.UUID | str | None,
) -> PaymentRecord:
    """Create exactly one pending payment record in the caller's transaction.

    Final approval uses this helper before its own commit, keeping the report,
    payment row, event, and audit log atomic.  The database unique constraint
    is the final protection against concurrent final-approval requests.
    """

    existing = (
        db.query(PaymentRecord)
        .filter(
            PaymentRecord.expense_report_id == report.id,
            PaymentRecord.is_deleted.is_(False),
        )
        .first()
    )
    if existing is not None:
        return existing

    payment = PaymentRecord(
        expense_report_id=report.id,
        payment_reference=f"PAY-{report.report_number}",
        amount=_money(report.total_amount),
        status="pending",
    )
    db.add(payment)
    db.flush()
    _event(
        db,
        payment,
        event_type="created",
        from_status=None,
        to_status="pending",
        performed_by=created_by,
        remarks="Created when the report completed approval.",
    )
    record_audit(
        db,
        "payment_records",
        str(payment.id),
        "created",
        after=_payment_audit_payload(payment),
        performed_by=str(created_by) if created_by else None,
    )
    return payment


def list_payments(
    db: Session,
    organization_id: uuid.UUID | str,
    *,
    status: str | None = None,
    batch_id: uuid.UUID | str | None = None,
    limit: int = 50,
    offset: int = 0,
) -> tuple[list[PaymentRecord], int]:
    organization = _as_uuid(organization_id, field_name="organization id")
    if status is not None and status not in PAYMENT_STATUSES:
        raise PaymentValidationError("Unsupported payment status")
    if limit < 1 or limit > 200:
        raise PaymentValidationError("limit must be between 1 and 200")
    if offset < 0:
        raise PaymentValidationError("offset must be zero or greater")

    query = (
        db.query(PaymentRecord)
        .join(ExpenseReport, ExpenseReport.id == PaymentRecord.expense_report_id)
        .join(User, User.id == ExpenseReport.employee_user_id)
        .filter(
            PaymentRecord.is_deleted.is_(False),
            ExpenseReport.is_deleted.is_(False),
            User.is_deleted.is_(False),
            User.organization_id == organization,
        )
    )
    if status is not None:
        query = query.filter(PaymentRecord.status == status)
    if batch_id is not None:
        query = query.filter(PaymentRecord.payment_batch_id == _as_uuid(batch_id, field_name="payment batch id"))
    total = query.count()
    payments = query.order_by(PaymentRecord.created_at.desc()).offset(offset).limit(limit).all()
    return payments, total


def get_payment(
    db: Session,
    payment_id: uuid.UUID | str,
    organization_id: uuid.UUID | str,
    *,
    for_update: bool = False,
) -> PaymentRecord:
    organization = _as_uuid(organization_id, field_name="organization id")
    payment_uuid = _as_uuid(payment_id, field_name="payment id")
    query = _scoped_payment_query(db, payment_uuid, organization)
    if for_update:
        query = query.with_for_update()
    payment = query.first()
    if payment is None:
        raise PaymentNotFoundError("Payment not found")
    return payment


def _get_batch(
    db: Session,
    batch_id: uuid.UUID | str,
    organization_id: uuid.UUID | str,
    *,
    for_update: bool = False,
) -> PaymentBatch:
    organization = _as_uuid(organization_id, field_name="organization id")
    query = db.query(PaymentBatch).filter(
        PaymentBatch.id == _as_uuid(batch_id, field_name="payment batch id"),
        PaymentBatch.organization_id == organization,
        PaymentBatch.is_deleted.is_(False),
    )
    if for_update:
        query = query.with_for_update()
    batch = query.first()
    if batch is None:
        raise PaymentNotFoundError("Payment batch not found")
    return batch


def list_batches(
    db: Session,
    organization_id: uuid.UUID | str,
    *,
    limit: int = 50,
    offset: int = 0,
) -> tuple[list[PaymentBatch], int]:
    organization = _as_uuid(organization_id, field_name="organization id")
    if limit < 1 or limit > 200:
        raise PaymentValidationError("limit must be between 1 and 200")
    if offset < 0:
        raise PaymentValidationError("offset must be zero or greater")
    query = db.query(PaymentBatch).filter(
        PaymentBatch.organization_id == organization,
        PaymentBatch.is_deleted.is_(False),
    )
    return (
        query.order_by(PaymentBatch.created_at.desc()).offset(offset).limit(limit).all(),
        query.count(),
    )


def get_batch(db: Session, batch_id: uuid.UUID | str, organization_id: uuid.UUID | str) -> PaymentBatch:
    return _get_batch(db, batch_id, organization_id)


def create_batch(
    db: Session,
    organization_id: uuid.UUID | str,
    actor_user_id: uuid.UUID | str,
    payment_ids: Iterable[uuid.UUID | str],
    *,
    remarks: str | None = None,
) -> PaymentBatch:
    """Move selected pending payments into one single-currency export batch."""

    organization = _as_uuid(organization_id, field_name="organization id")
    actor_id = _require_actor(db, actor_user_id, organization)
    resolved_ids = [_as_uuid(payment_id, field_name="payment id") for payment_id in payment_ids]
    if not resolved_ids:
        raise PaymentValidationError("Select at least one pending payment")
    if len(set(resolved_ids)) != len(resolved_ids):
        raise PaymentValidationError("Payment IDs must not contain duplicates")

    rows = (
        db.query(PaymentRecord, ExpenseReport)
        .join(ExpenseReport, ExpenseReport.id == PaymentRecord.expense_report_id)
        .join(User, User.id == ExpenseReport.employee_user_id)
        .filter(
            PaymentRecord.id.in_(resolved_ids),
            PaymentRecord.is_deleted.is_(False),
            ExpenseReport.is_deleted.is_(False),
            User.is_deleted.is_(False),
            User.organization_id == organization,
        )
        .with_for_update()
        .all()
    )
    if len(rows) != len(resolved_ids):
        raise PaymentNotFoundError("One or more payments were not found")
    invalid = [payment.payment_reference for payment, _report in rows if payment.status != "pending"]
    if invalid:
        raise PaymentTransitionError("Only pending payments can be added to a batch")

    currencies = {report.currency_code.upper() for _payment, report in rows}
    if len(currencies) != 1:
        raise PaymentValidationError("A payment batch must contain one currency")

    total = sum((_money(payment.amount) for payment, _report in rows), Decimal("0.00"))
    batch = PaymentBatch(
        organization_id=organization,
        batch_reference=f"PB-{uuid.uuid4().hex[:12].upper()}",
        status="created",
        currency_code=currencies.pop(),
        total_amount=total,
        payment_count=len(rows),
        created_by=actor_id,
        remarks=(remarks or "").strip() or None,
    )
    db.add(batch)
    db.flush()
    for payment, _report in rows:
        before = _payment_audit_payload(payment)
        payment.payment_batch_id = batch.id
        payment.status = "batched"
        _event(
            db,
            payment,
            event_type="batched",
            from_status="pending",
            to_status="batched",
            performed_by=actor_id,
            remarks=batch.remarks,
        )
        record_audit(
            db,
            "payment_records",
            str(payment.id),
            "batched",
            before=before,
            after=_payment_audit_payload(payment),
            performed_by=str(actor_id),
        )
    record_audit(
        db,
        "payment_batches",
        str(batch.id),
        "created",
        after={
            "batch_reference": batch.batch_reference,
            "status": batch.status,
            "payment_count": batch.payment_count,
            "total_amount": str(batch.total_amount),
            "currency": batch.currency_code,
        },
        performed_by=str(actor_id),
    )
    db.commit()
    db.refresh(batch)
    return batch


def _batch_rows_for_export(db: Session, batch: PaymentBatch):
    return (
        db.query(PaymentRecord, ExpenseReport, User)
        .join(ExpenseReport, ExpenseReport.id == PaymentRecord.expense_report_id)
        .join(User, User.id == ExpenseReport.employee_user_id)
        .filter(
            PaymentRecord.payment_batch_id == batch.id,
            PaymentRecord.is_deleted.is_(False),
            ExpenseReport.is_deleted.is_(False),
            User.is_deleted.is_(False),
        )
        .order_by(PaymentRecord.payment_reference)
        .with_for_update()
        .all()
    )


def _csv_for_batch(batch: PaymentBatch, rows) -> str:
    """Generate export content only from approved report/payment identifiers.

    Recipient bank data intentionally does not appear in either CSV exports or
    persistence.  An external payment integration can map the employee number
    to a payment recipient in its own secure domain.
    """

    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(
        [
            "batch_reference",
            "payment_reference",
            "report_number",
            "employee_number",
            "amount",
            "currency",
        ]
    )
    for payment, report, employee in rows:
        writer.writerow(
            [
                batch.batch_reference,
                payment.payment_reference,
                report.report_number,
                employee.employee_number,
                f"{_money(payment.amount):.2f}",
                report.currency_code,
            ]
        )
    return output.getvalue()


def _export_values(batch: PaymentBatch, rows) -> tuple[list[str], list[list[str]]]:
    headers = ["Batch reference", "Payment reference", "Report number", "Employee number", "Amount", "Currency"]
    values = [
        [
            batch.batch_reference,
            payment.payment_reference,
            report.report_number,
            employee.employee_number,
            f"{_money(payment.amount):.2f}",
            report.currency_code,
        ]
        for payment, report, employee in rows
    ]
    return headers, values


def _xlsx_for_batch(batch: PaymentBatch, rows) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reimbursements"
    headers, values = _export_values(batch, rows)
    sheet.append(["Presidio reimbursement export", batch.batch_reference])
    sheet.append(["Currency", batch.currency_code])
    sheet.append(["Total", f"{_money(batch.total_amount):.2f}"])
    sheet.append([])
    sheet.append(headers)
    for cell in sheet[5]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="202020")
    for row in values:
        sheet.append(row)
    sheet.freeze_panes = "A6"
    for column, width in {"A": 22, "B": 24, "C": 18, "D": 20, "E": 16, "F": 12}.items():
        sheet.column_dimensions[column].width = width
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_for_batch(batch: PaymentBatch, rows) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = io.BytesIO()
    document = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=28, rightMargin=28, topMargin=28, bottomMargin=28)
    styles = getSampleStyleSheet()
    headers, values = _export_values(batch, rows)
    elements = [
        Paragraph("Presidio reimbursement export", styles["Title"]),
        Paragraph(f"Batch {batch.batch_reference} · {batch.payment_count} reimbursements · {batch.currency_code} {_money(batch.total_amount):.2f}", styles["Normal"]),
        Spacer(1, 14),
    ]
    table = Table([headers, *values], repeatRows=1, colWidths=[110, 125, 100, 110, 80, 65])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#202020")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BBBBBB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F0E8")]),
        ("ALIGN", (4, 1), (4, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    elements.append(table)
    document.build(elements)
    return output.getvalue()


def export_batch_file(
    db: Session,
    batch_id: uuid.UUID | str,
    organization_id: uuid.UUID | str,
    actor_user_id: uuid.UUID | str,
    *,
    file_format: str,
) -> tuple[PaymentBatch, bytes | str, str, str]:
    """Export an auditable payment batch in CSV, Excel, or PDF form."""

    if file_format not in {"csv", "xlsx", "pdf"}:
        raise PaymentValidationError("Unsupported export format")
    batch, csv_content = export_batch(db, batch_id, organization_id, actor_user_id)
    rows = _batch_rows_for_export(db, batch)
    if file_format == "csv":
        return batch, csv_content, "text/csv; charset=utf-8", "csv"
    if file_format == "xlsx":
        return batch, _xlsx_for_batch(batch, rows), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
    return batch, _pdf_for_batch(batch, rows), "application/pdf", "pdf"


def export_batch(
    db: Session,
    batch_id: uuid.UUID | str,
    organization_id: uuid.UUID | str,
    actor_user_id: uuid.UUID | str,
) -> tuple[PaymentBatch, str]:
    """Produce a CSV and atomically mark a created batch as exported.

    Re-downloading an already-exported batch is deliberately idempotent and
    does not create duplicate lifecycle events.
    """

    organization = _as_uuid(organization_id, field_name="organization id")
    actor_id = _require_actor(db, actor_user_id, organization)
    batch = _get_batch(db, batch_id, organization, for_update=True)
    if batch.status not in BATCH_STATUSES:
        raise PaymentTransitionError("Only created or exported payment batches can be exported")
    rows = _batch_rows_for_export(db, batch)
    if len(rows) != batch.payment_count:
        raise PaymentTransitionError("Payment batch contents no longer match its recorded count")

    csv_content = _csv_for_batch(batch, rows)
    if batch.status == "exported":
        return batch, csv_content

    if any(payment.status != "batched" for payment, _report, _employee in rows):
        raise PaymentTransitionError("Only batched payments can be exported")

    exported_at = utcnow()
    batch.status = "exported"
    batch.exported_at = exported_at
    for payment, _report, _employee in rows:
        before = _payment_audit_payload(payment)
        payment.status = "exported"
        payment.exported_at = exported_at
        _event(
            db,
            payment,
            event_type="exported",
            from_status="batched",
            to_status="exported",
            performed_by=actor_id,
            remarks=f"Exported in batch {batch.batch_reference}.",
        )
        record_audit(
            db,
            "payment_records",
            str(payment.id),
            "exported",
            before=before,
            after=_payment_audit_payload(payment),
            performed_by=str(actor_id),
        )
    record_audit(
        db,
        "payment_batches",
        str(batch.id),
        "exported",
        before={"status": "created"},
        after={"status": batch.status, "exported_at": exported_at.isoformat()},
        performed_by=str(actor_id),
    )
    db.commit()
    db.refresh(batch)
    return batch, csv_content


def mark_paid(
    db: Session,
    payment_id: uuid.UUID | str,
    organization_id: uuid.UUID | str,
    actor_user_id: uuid.UUID | str,
    *,
    provider_reference: str,
    payment_date: date | None = None,
    remarks: str | None = None,
) -> PaymentRecord:
    """Complete an exported payment and publish employee-visible status."""

    organization = _as_uuid(organization_id, field_name="organization id")
    actor_id = _require_actor(db, actor_user_id, organization)
    reference = provider_reference.strip()
    if not reference:
        raise PaymentValidationError("A payment provider reference is required")
    payment = get_payment(db, payment_id, organization, for_update=True)
    if payment.status != "exported":
        raise PaymentTransitionError("Only exported payments can be marked paid")

    report = _payment_report(db, payment)
    before = _payment_audit_payload(payment)
    paid_on = payment_date or date.today()
    payment.status = "paid"
    payment.payment_date = paid_on
    payment.provider_reference = reference
    payment.failure_reason = None
    payment.processed_by = actor_id
    payment.remarks = (remarks or "").strip() or None
    _event(
        db,
        payment,
        event_type="paid",
        from_status="exported",
        to_status="paid",
        performed_by=actor_id,
        remarks=payment.remarks,
        provider_reference=reference,
    )
    record_audit(
        db,
        "payment_records",
        str(payment.id),
        "paid",
        before=before,
        after=_payment_audit_payload(payment),
        performed_by=str(actor_id),
    )
    if report.status == "approved_pending_payment":
        report_before = {"status": report.status}
        report.status = "paid"
        record_audit(
            db,
            "expense_reports",
            str(report.id),
            "payment_paid",
            before=report_before,
            after={"status": report.status, "payment_id": str(payment.id)},
            performed_by=str(actor_id),
        )
    notify(
        db,
        report.employee_user_id,
        "payment_paid",
        {
            "title": "Reimbursement paid",
            "body": f"Your reimbursement for {report.title} has been paid.",
            "report_id": str(report.id),
            "type": "payment_status",
        },
        channels=("in_app", "email"),
    )
    db.commit()
    db.refresh(payment)
    return payment


def mark_failed(
    db: Session,
    payment_id: uuid.UUID | str,
    organization_id: uuid.UUID | str,
    actor_user_id: uuid.UUID | str,
    *,
    failure_reason: str,
    remarks: str | None = None,
) -> PaymentRecord:
    """Record a finance-visible failure without changing approval outcome."""

    organization = _as_uuid(organization_id, field_name="organization id")
    actor_id = _require_actor(db, actor_user_id, organization)
    reason = failure_reason.strip()
    if not reason:
        raise PaymentValidationError("A payment failure reason is required")
    payment = get_payment(db, payment_id, organization, for_update=True)
    # A batched payment is still part of a mutable, un-exported CSV batch.
    # Marking it failed at this point would leave the batch's persisted count
    # and total out of sync, and make that batch impossible to export. A
    # finance failure therefore records a payment-rail outcome only after the
    # batch has been exported.
    if payment.status != "exported":
        raise PaymentTransitionError("Only exported payments can be marked failed")

    report = _payment_report(db, payment)
    before = _payment_audit_payload(payment)
    from_status = payment.status
    payment.status = "failed"
    payment.failure_reason = reason
    payment.processed_by = actor_id
    payment.remarks = (remarks or "").strip() or None
    _event(
        db,
        payment,
        event_type="failed",
        from_status=from_status,
        to_status="failed",
        performed_by=actor_id,
        remarks=payment.remarks or reason,
    )
    record_audit(
        db,
        "payment_records",
        str(payment.id),
        "failed",
        before=before,
        after=_payment_audit_payload(payment),
        performed_by=str(actor_id),
    )
    notify(
        db,
        report.employee_user_id,
        "payment_failed",
        {
            "title": "Reimbursement payment needs attention",
            "body": f"Finance is reviewing the payment for {report.title}.",
            "report_id": str(report.id),
            "type": "payment_status",
        },
        channels=("in_app", "email"),
    )
    db.commit()
    db.refresh(payment)
    return payment


def payment_event_payload(db: Session, event: PaymentEvent) -> dict[str, Any]:
    actor = db.get(User, event.performed_by) if event.performed_by else None
    return {
        "id": str(event.id),
        "event_type": event.event_type,
        "from_status": event.from_status,
        "to_status": event.to_status,
        "amount": float(event.amount),
        "provider_reference": event.provider_reference,
        "remarks": event.remarks,
        "performed_by": str(event.performed_by) if event.performed_by else None,
        "performed_by_name": actor.full_name if actor else None,
        "occurred_at": event.occurred_at.isoformat() if event.occurred_at else None,
    }


def payment_payload(
    db: Session,
    payment: PaymentRecord,
    *,
    include_history: bool = False,
    include_finance_details: bool = False,
) -> dict[str, Any]:
    """Serialize payment state without ever returning a bank-detail field."""

    report = _payment_report(db, payment)
    employee = db.get(User, report.employee_user_id)
    batch = db.get(PaymentBatch, payment.payment_batch_id) if payment.payment_batch_id else None
    payload: dict[str, Any] = {
        "id": str(payment.id),
        "report_id": str(report.id),
        "report_number": report.report_number,
        "employee_name": employee.full_name if employee else None,
        "employee_number": employee.employee_number if employee else None,
        "payment_reference": payment.payment_reference,
        "amount": float(payment.amount),
        "currency": report.currency_code,
        "status": payment.status,
        "payment_date": payment.payment_date.isoformat() if payment.payment_date else None,
        "exported_at": payment.exported_at.isoformat() if payment.exported_at else None,
        "batch": (
            {
                "id": str(batch.id),
                "batch_reference": batch.batch_reference,
                "status": batch.status,
            }
            if batch is not None and not batch.is_deleted
            else None
        ),
    }
    if include_finance_details:
        payload["provider_reference"] = payment.provider_reference
        payload["failure_reason"] = payment.failure_reason
        payload["remarks"] = payment.remarks
        payload["processed_by"] = str(payment.processed_by) if payment.processed_by else None
    if include_history:
        events = (
            db.query(PaymentEvent)
            .filter(
                PaymentEvent.payment_record_id == payment.id,
                PaymentEvent.is_deleted.is_(False),
            )
            .order_by(PaymentEvent.occurred_at, PaymentEvent.created_at)
            .all()
        )
        payload["history"] = [payment_event_payload(db, event) for event in events]
    return payload


def payment_for_report_payload(db: Session, report: ExpenseReport) -> dict[str, Any] | None:
    payment = (
        db.query(PaymentRecord)
        .filter(
            PaymentRecord.expense_report_id == report.id,
            PaymentRecord.is_deleted.is_(False),
        )
        .first()
    )
    return payment_payload(db, payment) if payment is not None else None


def batch_payload(db: Session, batch: PaymentBatch, *, include_payments: bool = False) -> dict[str, Any]:
    creator = db.get(User, batch.created_by)
    payload: dict[str, Any] = {
        "id": str(batch.id),
        "batch_reference": batch.batch_reference,
        "status": batch.status,
        "currency": batch.currency_code,
        "total_amount": float(batch.total_amount),
        "payment_count": batch.payment_count,
        "created_by": str(batch.created_by),
        "created_by_name": creator.full_name if creator else None,
        "created_at": batch.created_at.isoformat() if batch.created_at else None,
        "exported_at": batch.exported_at.isoformat() if batch.exported_at else None,
        "remarks": batch.remarks,
    }
    if include_payments:
        payments = (
            db.query(PaymentRecord)
            .filter(
                PaymentRecord.payment_batch_id == batch.id,
                PaymentRecord.is_deleted.is_(False),
            )
            .order_by(PaymentRecord.payment_reference)
            .all()
        )
        payload["payments"] = [payment_payload(db, payment, include_finance_details=True) for payment in payments]
    return payload
