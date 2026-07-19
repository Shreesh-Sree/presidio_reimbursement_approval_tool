"""Policy template generation and extraction service for Excel and PDF files."""

from __future__ import annotations

import io
import re
import csv
from decimal import Decimal
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pypdf
from sqlalchemy.orm import Session

from app.services import category_service, policy_service
from app.models.policy import Policy, PolicyRule
from app.models.expense_category import ExpenseCategory


def generate_excel_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Policy Rules Template"

    # Header styling
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    headers = [
        "Category Code",
        "Category Name",
        "Max Per Day ($)",
        "Max Per Trip ($)",
        "Category Cap ($)",
        "Receipt Required Above ($)",
        "Description",
    ]

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sample_rows = [
        ["MEALS", "Meals & Entertainment", 75.00, 300.00, 1000.00, 25.00, "Daily meal allowance for business travel"],
        ["TRAVEL", "Domestic Travel", 250.00, 1500.00, 5000.00, 50.00, "Flight, train, and lodging expense policy"],
        ["SUPPLIES", "Office Supplies", 100.00, 250.00, 800.00, 20.00, "Stationery and workstation equipment"],
        ["COMMUTE", "Taxi & Rideshare", 50.00, 200.00, 600.00, 15.00, "Local transit and rideshare reimbursement"],
    ]

    for row in sample_rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=len(sample_rows) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.number_format = "$#,##0.00"

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 26
    ws.column_dimensions["G"].width = 45

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pdf_template() -> bytes:
    """Generates a formatted PDF template containing structured policy instructions and sample rules."""
    output = io.BytesIO()
    
    content_lines = [
        "CORPORATE EXPENSE POLICY TEMPLATE",
        "=====================================",
        "",
        "Instructions: Use this format to structure policy rules for upload.",
        "",
        "Rule 1:",
        "Category: Meals & Entertainment",
        "Code: MEALS",
        "Max Per Day: $75.00",
        "Max Per Trip: $300.00",
        "Receipt Required Above: $25.00",
        "Description: Daily meal allowance for business trips.",
        "",
        "Rule 2:",
        "Category: Domestic Travel",
        "Code: TRAVEL",
        "Max Per Day: $250.00",
        "Max Per Trip: $1500.00",
        "Receipt Required Above: $50.00",
        "Description: Flight, train, and lodging reimbursement.",
        "",
        "Rule 3:",
        "Category: Office Supplies",
        "Code: SUPPLIES",
        "Max Per Day: $100.00",
        "Max Per Trip: $250.00",
        "Receipt Required Above: $20.00",
        "Description: Stationery and hardware equipment.",
    ]

    pdf_headers = "%PDF-1.4\n"
    body = (
        "1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n"
        "2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n"
        "3 0 obj\n<< /Type /Page /Parent 2 0 R /Resources 4 0 R /MediaBox [0 0 612 792] /Contents 5 0 R >>\nendobj\n"
        "4 0 obj\n<< /Font << /F1 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> >> >>\nendobj\n"
    )
    
    stream_data = "BT /F1 10 Tf 50 720 Td 14 TL\n"
    for line in content_lines:
        escaped = line.replace("(", "\\(").replace(")", "\\)")
        stream_data += f"({escaped}) '\n"
    stream_data += "ET"

    stream_obj = f"5 0 obj\n<< /Length {len(stream_data)} >>\nstream\n{stream_data}\nendstream\nendobj\n"
    full_pdf = (pdf_headers + body + stream_obj).encode("latin1")
    output.write(full_pdf)
    return output.getvalue()


def extract_rules_from_excel(content: bytes) -> list[dict[str, Any]]:
    rules = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return rules

        start_idx = 1
        for idx, row in enumerate(rows):
            if row and any("category" in str(cell).lower() for cell in row if cell):
                start_idx = idx + 1
                break

        for row in rows[start_idx:]:
            if not row or not any(row):
                continue
            code = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
            name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else code
            if not code or code.lower() in ("category code", "code"):
                continue

            def parse_num(val: Any) -> float | None:
                if val is None or val == "":
                    return None
                try:
                    cleaned = re.sub(r"[^\d\.]", "", str(val))
                    return float(cleaned) if cleaned else None
                except ValueError:
                    return None

            max_per_day = parse_num(row[2]) if len(row) > 2 else None
            max_per_trip = parse_num(row[3]) if len(row) > 3 else None
            per_category_cap = parse_num(row[4]) if len(row) > 4 else None
            receipt_required_above = parse_num(row[5]) if len(row) > 5 else None
            description = str(row[6]).strip() if len(row) > 6 and row[6] is not None else None

            rules.append({
                "code": code,
                "name": name,
                "max_per_day": max_per_day,
                "max_per_trip": max_per_trip,
                "per_category_cap": per_category_cap,
                "receipt_required_above": receipt_required_above,
                "description": description,
            })
    except Exception:
        pass
    return rules


def extract_rules_from_pdf(content: bytes) -> list[dict[str, Any]]:
    rules = []
    try:
        reader = pypdf.PdfReader(io.BytesIO(content))
        full_text = "\n".join([page.extract_text() or "" for page in reader.pages])

        blocks = re.split(r"(?:Rule\s*\d+:|Category:)", full_text, flags=re.IGNORECASE)
        for block in blocks:
            if not block.strip():
                continue
            
            category_match = re.search(r"(?:Category|Name):\s*([^\n]+)", block, re.IGNORECASE)
            code_match = re.search(r"Code:\s*([A-Za-z0-9_-]+)", block, re.IGNORECASE)
            day_match = re.search(r"Max\s*Per\s*Day:\s*\$?([\d\.]+)", block, re.IGNORECASE)
            trip_match = re.search(r"Max\s*Per\s*Trip:\s*\$?([\d\.]+)", block, re.IGNORECASE)
            receipt_match = re.search(r"Receipt\s*Required\s*Above:\s*\$?([\d\.]+)", block, re.IGNORECASE)
            desc_match = re.search(r"Description:\s*([^\n]+)", block, re.IGNORECASE)

            if category_match or code_match or day_match:
                name = category_match.group(1).strip() if category_match else "General"
                code = code_match.group(1).strip().upper() if code_match else re.sub(r"[^A-Z0-9]", "", name.upper())[:15]
                rules.append({
                    "code": code or "GENERAL",
                    "name": name,
                    "max_per_day": float(day_match.group(1)) if day_match else None,
                    "max_per_trip": float(trip_match.group(1)) if trip_match else None,
                    "per_category_cap": None,
                    "receipt_required_above": float(receipt_match.group(1)) if receipt_match else None,
                    "description": desc_match.group(1).strip() if desc_match else None,
                })
    except Exception:
        pass
    return rules


def apply_extracted_rules(
    db: Session,
    organization_id: str,
    rules: list[dict[str, Any]],
) -> dict[str, Any]:
    """Applies extracted rules by creating/updating categories and associating PolicyRule entries."""
    created_count = 0
    updated_count = 0

    for rule in rules:
        code = rule["code"]
        name = rule["name"]
        max_per_day = rule.get("max_per_day")
        max_per_trip = rule.get("max_per_trip")
        per_category_cap = rule.get("per_category_cap")
        receipt_required_above = rule.get("receipt_required_above")
        description = rule.get("description")

        existing_cat = category_service.get_category_by_code(db, code, organization_id)
        if existing_cat is None:
            category_service.create_category(
                db,
                code=code,
                name=name,
                organization_id=organization_id,
                description=description,
                max_per_day=max_per_day,
                max_per_trip=max_per_trip,
                per_category_cap=per_category_cap,
                receipt_required_above=receipt_required_above,
            )
            created_count += 1
        else:
            category_service.update_category(
                db,
                category_id=existing_cat.id,
                organization_id=organization_id,
                name=name,
                description=description,
                max_per_day=max_per_day,
                max_per_trip=max_per_trip,
                per_category_cap=per_category_cap,
                receipt_required_above=receipt_required_above,
            )
            updated_count += 1

    return {"status": "success", "created": created_count, "updated": updated_count, "total": len(rules)}
